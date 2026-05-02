import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime


def export_to_excel(results, coverage_report, selected_results):

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "Test Case ID",
        "Test Case Name",
        "Precondition",
        "Test Data",
        "Test Step",
        "Expected Result",
        "Actual Result",
        "Priority",
        "Pass/Fail",
        "Status",
        "Date",
        "Evidence"
    ]

    # =========================
    # STYLE HEADER
    # =========================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # =========================
    # DATA ROWS
    # =========================
    for tc in results:

        row = [
            tc["id"],
            tc["name"],
            tc["precondition"],
            json.dumps(tc["test_data"], ensure_ascii=False),
            tc["steps"],
            json.dumps(tc["expected"], ensure_ascii=False),
            str(tc.get("actual", "")),
            tc["priority"],
            tc.get("pass_fail", ""),
            tc.get("status", ""),
            tc["date"],
            ""
        ]

        ws.append(row)

    # =========================
    # AUTO FORMAT COLUMNS
    # =========================
    column_widths = [15, 25, 20, 40, 50, 40, 15, 10, 12, 10, 15, 20]

    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # wrap text + align all cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    # freeze header
    ws.freeze_panes = "A2"

    # save
    wb.save("test_report.xlsx")

    print("✅ Exported QA Test Report (Formatted)")